"""Testes da execução de simulações via API."""

import pytest
from fastapi.testclient import TestClient

CSV = "text/csv"


def _importar(client: TestClient, rota: str, conteudo: bytes) -> None:
    resposta = client.post(rota, files={"arquivo": ("dados.csv", conteudo, CSV)})
    assert resposta.status_code == 200, resposta.json()


@pytest.fixture
def base_com_instrutor(client: TestClient) -> TestClient:
    """Um instrutor multi-tipologia, sem turmas em andamento, pronto para simular."""
    from tests.fabricas import planilha_instrutores, planilha_tipologias

    _importar(
        client,
        "/importar/instrutores",
        planilha_instrutores(
            [["Maria Silva", "Jovem Digital", "manha_1", "2;3;4;5", "Programação;Pixel Art"]]
        ),
    )
    _importar(
        client,
        "/importar/tipologias",
        planilha_tipologias([["Programação", "60", "3"], ["Pixel Art", "24", "2"]]),
    )
    return client


@pytest.fixture
def cenario_basico(base_com_instrutor: TestClient) -> dict:
    resposta = base_com_instrutor.post(
        "/cenarios",
        json={
            "nome": "Cenário de teste",
            "periodo_de": "2026-08-31",
            "periodo_ate": "2026-11-30",
            "pesos_objetivo": {
                "maximizar_aproveitamento": 1.0,
                "antecipar_inicio": 0.0,
                "balancear_carga_instrutores": 0.0,
                "balancear_tipologias": 0.0,
            },
        },
    )
    assert resposta.status_code == 201
    return resposta.json()


class TestDisparo:
    def test_retorna_pendente_antes_de_expor_o_resultado(
        self, base_com_instrutor: TestClient, cenario_basico: dict
    ) -> None:
        """A resposta reflete o estado no momento do disparo, não após a
        conclusão em segundo plano."""
        resposta = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario_basico["id"]}
        )

        assert resposta.status_code == 202
        assert resposta.json()["status"] == "pendente"

    def test_simulacao_conclui_e_fica_consultavel(
        self, base_com_instrutor: TestClient, cenario_basico: dict
    ) -> None:
        disparo = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario_basico["id"]}
        ).json()

        consulta = base_com_instrutor.get(f"/simulacoes/{disparo['id']}").json()

        assert consulta["status"] == "concluida"
        assert consulta["solver_status"] in ("OTIMO", "FACTIVEL")
        assert consulta["tempo_execucao_seg"] is not None

    def test_cenario_inexistente_retorna_404(self, base_com_instrutor: TestClient) -> None:
        resposta = base_com_instrutor.post("/simulacoes/executar", json={"cenario_id": 9999})
        assert resposta.status_code == 404

    def test_bloqueia_por_tipologia_pendente(self, base_com_instrutor: TestClient) -> None:
        from tests.fabricas import planilha_instrutores

        # Importa um instrutor apto a uma tipologia inédita, ainda sem carga horária.
        _importar(
            base_com_instrutor,
            "/importar/instrutores",
            planilha_instrutores(
                [["João Souza", "Jovem Digital", "noite", "2;4", "Robótica"]]
            ),
        )
        cenario = base_com_instrutor.post(
            "/cenarios",
            json={
                "nome": "Com pendência",
                "periodo_de": "2026-08-31",
                "periodo_ate": "2026-11-30",
                "pesos_objetivo": {
                    "maximizar_aproveitamento": 1.0,
                    "antecipar_inicio": 0.0,
                    "balancear_carga_instrutores": 0.0,
                    "balancear_tipologias": 0.0,
                },
            },
        ).json()

        resposta = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario["id"]}
        )

        assert resposta.status_code == 422
        assert "Robótica" in resposta.json()["detail"]

    def test_bloqueia_escopo_sem_instrutores(self, client: TestClient) -> None:
        projeto = client.post("/projetos", json={"nome": "Projeto Vazio"}).json()
        cenario = client.post(
            "/cenarios",
            json={
                "nome": "Escopo vazio",
                "periodo_de": "2026-08-31",
                "periodo_ate": "2026-11-30",
                "projeto_ids": [projeto["id"]],
                "pesos_objetivo": {
                    "maximizar_aproveitamento": 1.0,
                    "antecipar_inicio": 0.0,
                    "balancear_carga_instrutores": 0.0,
                    "balancear_tipologias": 0.0,
                },
            },
        ).json()

        resposta = client.post("/simulacoes/executar", json={"cenario_id": cenario["id"]})

        assert resposta.status_code == 422
        assert "instrutor" in resposta.json()["detail"].lower()


class TestPersistenciaDoResultado:
    def test_turmas_sugeridas_sao_persistidas_com_encontros(
        self, base_com_instrutor: TestClient, cenario_basico: dict
    ) -> None:
        disparo = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario_basico["id"]}
        ).json()

        turmas = base_com_instrutor.get(f"/simulacoes/{disparo['id']}/turmas-sugeridas").json()

        assert len(turmas) > 0
        primeira = turmas[0]
        assert primeira["instrutor_nome"] == "Maria Silva"
        assert primeira["tipologia_nome"] in ("Programação", "Pixel Art")
        assert len(primeira["encontros"]) == primeira["num_encontros"]

    def test_kpis_sao_persistidos(
        self, base_com_instrutor: TestClient, cenario_basico: dict
    ) -> None:
        disparo = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario_basico["id"]}
        ).json()

        kpis = base_com_instrutor.get(f"/simulacoes/{disparo['id']}/kpis").json()

        assert kpis["total_turmas_sugeridas"] > 0
        assert 0 <= kpis["pct_ociosidade"] <= 100

    def test_resultado_vazio_e_conclusao_normal_nao_erro(self, client: TestClient) -> None:
        """Sem nenhum instrutor apto a nada gerar candidata, a simulação
        ainda deve concluir normalmente — só que sem nenhuma turma."""
        from tests.fabricas import planilha_instrutores, planilha_tipologias

        _importar(
            client,
            "/importar/instrutores",
            planilha_instrutores(
                [["Maria Silva", "Jovem Digital", "manha_1", "2;4", "Programação"]]
            ),
        )
        _importar(
            client,
            "/importar/tipologias",
            # Carga horária inviável para o período (curtíssimo) força zero candidatas.
            planilha_tipologias([["Programação", "60", "3"]]),
        )
        cenario = client.post(
            "/cenarios",
            json={
                "nome": "Período impossível",
                "periodo_de": "2026-08-31",
                "periodo_ate": "2026-09-02",  # curto demais para qualquer turma
                "pesos_objetivo": {
                    "maximizar_aproveitamento": 1.0,
                    "antecipar_inicio": 0.0,
                    "balancear_carga_instrutores": 0.0,
                    "balancear_tipologias": 0.0,
                },
            },
        ).json()

        disparo = client.post("/simulacoes/executar", json={"cenario_id": cenario["id"]}).json()
        consulta = client.get(f"/simulacoes/{disparo['id']}").json()

        assert consulta["status"] == "concluida"
        assert consulta["mensagem_erro"] is None

        turmas = client.get(f"/simulacoes/{disparo['id']}/turmas-sugeridas").json()
        assert turmas == []


class TestOportunidades:
    def test_reflete_o_leque_completo_nao_so_o_selecionado(
        self, base_com_instrutor: TestClient, cenario_basico: dict
    ) -> None:
        """Maria domina Programação e Pixel Art — ambas devem aparecer no
        mapa, mesmo que o solver só tenha aberto uma delas por data."""
        disparo = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario_basico["id"]}
        ).json()

        oportunidades = base_com_instrutor.get(f"/simulacoes/{disparo['id']}/oportunidades").json()

        tipologias_no_leque = {o["tipologia_nome"] for o in oportunidades}
        assert tipologias_no_leque == {"Programação", "Pixel Art"}

    def test_ordenado_cronologicamente(
        self, base_com_instrutor: TestClient, cenario_basico: dict
    ) -> None:
        disparo = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario_basico["id"]}
        ).json()

        oportunidades = base_com_instrutor.get(f"/simulacoes/{disparo['id']}/oportunidades").json()

        datas = [o["data_inicio"] for o in oportunidades]
        assert datas == sorted(datas)

    def test_filtra_por_tipologia(
        self, base_com_instrutor: TestClient, cenario_basico: dict
    ) -> None:
        disparo = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario_basico["id"]}
        ).json()
        tipologia_id = next(
            o["tipologia_id"]
            for o in base_com_instrutor.get(f"/simulacoes/{disparo['id']}/oportunidades").json()
        )

        filtradas = base_com_instrutor.get(
            f"/simulacoes/{disparo['id']}/oportunidades?tipologia_id={tipologia_id}"
        ).json()

        assert all(o["tipologia_id"] == tipologia_id for o in filtradas)


class TestAgenda:
    def test_combina_andamento_e_sugeridas(
        self, base_com_instrutor: TestClient, cenario_basico: dict
    ) -> None:
        instrutor_id = base_com_instrutor.get("/instrutores").json()[0]["id"]
        disparo = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario_basico["id"]}
        ).json()

        agenda = base_com_instrutor.get(f"/simulacoes/{disparo['id']}/agenda/{instrutor_id}").json()

        assert len(agenda) > 0
        assert all(item["origem"] == "sugerida" for item in agenda)  # sem turma em andamento

    def test_instrutor_inexistente_retorna_404(
        self, base_com_instrutor: TestClient, cenario_basico: dict
    ) -> None:
        disparo = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario_basico["id"]}
        ).json()

        resposta = base_com_instrutor.get(f"/simulacoes/{disparo['id']}/agenda/9999")

        assert resposta.status_code == 404


class TestCapacidadeInstrutores:
    def test_lista_utilizacao_de_cada_instrutor(
        self, base_com_instrutor: TestClient, cenario_basico: dict
    ) -> None:
        disparo = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario_basico["id"]}
        ).json()

        capacidade = base_com_instrutor.get(
            f"/simulacoes/{disparo['id']}/capacidade-instrutores"
        ).json()

        assert len(capacidade) == 1
        item = capacidade[0]
        assert item["instrutor_nome"] == "Maria Silva"
        assert item["slots_disponiveis"] > 0
        assert 0 <= item["utilizacao_percentual"] <= 100
        assert item["primeira_data_livre"] is not None
        assert "manha_1" in item["primeira_data_livre_por_slot"]

    def test_filtra_por_projeto(self, base_com_instrutor: TestClient, cenario_basico: dict) -> None:
        disparo = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario_basico["id"]}
        ).json()
        outro_projeto = base_com_instrutor.post("/projetos", json={"nome": "Outro Projeto"}).json()

        vazio = base_com_instrutor.get(
            f"/simulacoes/{disparo['id']}/capacidade-instrutores?projeto_id={outro_projeto['id']}"
        ).json()

        assert vazio == []

    def test_simulacao_inexistente_retorna_404(self, client: TestClient) -> None:
        assert client.get("/simulacoes/9999/capacidade-instrutores").status_code == 404


class TestReprodutibilidade:
    def test_editar_cenario_nao_altera_simulacao_ja_executada(
        self, base_com_instrutor: TestClient, cenario_basico: dict
    ) -> None:
        disparo = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario_basico["id"]}
        ).json()
        objetivo_original = base_com_instrutor.get(f"/simulacoes/{disparo['id']}").json()[
            "objetivo_valor"
        ]
        turmas_originais = base_com_instrutor.get(
            f"/simulacoes/{disparo['id']}/turmas-sugeridas"
        ).json()

        # Edita o cenário com pesos completamente diferentes.
        base_com_instrutor.put(
            f"/cenarios/{cenario_basico['id']}",
            json={
                "nome": cenario_basico["nome"],
                "periodo_de": cenario_basico["periodo_de"],
                "periodo_ate": cenario_basico["periodo_ate"],
                "pesos_objetivo": {
                    "maximizar_aproveitamento": 0.0,
                    "antecipar_inicio": 1.0,
                    "balancear_carga_instrutores": 0.0,
                    "balancear_tipologias": 0.0,
                },
            },
        )

        relido = base_com_instrutor.get(f"/simulacoes/{disparo['id']}").json()
        turmas_relidas = base_com_instrutor.get(
            f"/simulacoes/{disparo['id']}/turmas-sugeridas"
        ).json()

        assert relido["objetivo_valor"] == objetivo_original
        assert turmas_relidas == turmas_originais

    def test_mudar_turmas_em_andamento_nao_altera_simulacao_antiga(
        self, base_com_instrutor: TestClient, cenario_basico: dict
    ) -> None:
        """O snapshot de capacidade é congelado no momento da execução."""
        disparo = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario_basico["id"]}
        ).json()
        kpis_originais = base_com_instrutor.get(f"/simulacoes/{disparo['id']}/kpis").json()

        instrutor_id = base_com_instrutor.get("/instrutores").json()[0]["id"]
        tipologia_id = base_com_instrutor.get("/tipologias").json()[0]["id"]
        base_com_instrutor.post(
            "/turmas-em-andamento",
            json={
                "instrutor_id": instrutor_id,
                "tipologia_id": tipologia_id,
                "modalidade": "regular_seg_qua",
                "turno": "manha_1",
                "data_inicio": "2026-09-01",
                "data_fim_prevista": "2026-10-01",
            },
        )

        kpis_relidos = base_com_instrutor.get(f"/simulacoes/{disparo['id']}/kpis").json()

        assert kpis_relidos == kpis_originais


class TestComparacao:
    def test_compara_duas_simulacoes(
        self, base_com_instrutor: TestClient, cenario_basico: dict
    ) -> None:
        d1 = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario_basico["id"]}
        ).json()
        d2 = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario_basico["id"]}
        ).json()

        resposta = base_com_instrutor.get(f"/simulacoes/comparar?ids={d1['id']},{d2['id']}")

        assert resposta.status_code == 200
        corpo = resposta.json()
        assert len(corpo["itens"]) == 2
        assert corpo["periodos_divergentes"] is False

    def test_inclui_pesos_do_cenario_em_cada_item(
        self, base_com_instrutor: TestClient, cenario_basico: dict
    ) -> None:
        d1 = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario_basico["id"]}
        ).json()
        d2 = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario_basico["id"]}
        ).json()

        corpo = base_com_instrutor.get(f"/simulacoes/comparar?ids={d1['id']},{d2['id']}").json()

        assert corpo["itens"][0]["pesos_objetivo"]["maximizar_aproveitamento"] == 1.0

    def test_sinaliza_periodos_divergentes(self, base_com_instrutor: TestClient) -> None:
        cenario_curto = base_com_instrutor.post(
            "/cenarios",
            json={
                "nome": "Curto",
                "periodo_de": "2026-08-31",
                "periodo_ate": "2026-09-30",
                "pesos_objetivo": {
                    "maximizar_aproveitamento": 1.0,
                    "antecipar_inicio": 0.0,
                    "balancear_carga_instrutores": 0.0,
                    "balancear_tipologias": 0.0,
                },
            },
        ).json()
        cenario_longo = base_com_instrutor.post(
            "/cenarios",
            json={
                "nome": "Longo",
                "periodo_de": "2026-08-31",
                "periodo_ate": "2027-04-30",
                "pesos_objetivo": {
                    "maximizar_aproveitamento": 1.0,
                    "antecipar_inicio": 0.0,
                    "balancear_carga_instrutores": 0.0,
                    "balancear_tipologias": 0.0,
                },
            },
        ).json()
        d1 = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario_curto["id"]}
        ).json()
        d2 = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario_longo["id"]}
        ).json()

        corpo = base_com_instrutor.get(f"/simulacoes/comparar?ids={d1['id']},{d2['id']}").json()

        assert corpo["periodos_divergentes"] is True

    def test_recusa_identificador_inexistente(self, base_com_instrutor: TestClient) -> None:
        resposta = base_com_instrutor.get("/simulacoes/comparar?ids=9999,8888")
        assert resposta.status_code == 404

    def test_recusa_simulacao_nao_concluida(
        self, base_com_instrutor: TestClient, cenario_basico: dict, monkeypatch
    ) -> None:
        import app.api.simulacoes as modulo_simulacoes

        d1 = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario_basico["id"]}
        ).json()

        monkeypatch.setattr(modulo_simulacoes, "executar_simulacao", lambda _id: None)
        d2 = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario_basico["id"]}
        ).json()

        resposta = base_com_instrutor.get(f"/simulacoes/comparar?ids={d1['id']},{d2['id']}")

        assert resposta.status_code == 422
        assert str(d2["id"]) in resposta.json()["detail"]

    def test_recusa_menos_de_duas_simulacoes(
        self, base_com_instrutor: TestClient, cenario_basico: dict
    ) -> None:
        d1 = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario_basico["id"]}
        ).json()

        resposta = base_com_instrutor.get(f"/simulacoes/comparar?ids={d1['id']}")

        assert resposta.status_code == 422

    def test_rota_comparar_nao_colide_com_id_numerico(
        self, base_com_instrutor: TestClient, cenario_basico: dict
    ) -> None:
        """Garante que /simulacoes/comparar não seja capturada por /{simulacao_id}."""
        d1 = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario_basico["id"]}
        ).json()
        d2 = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario_basico["id"]}
        ).json()

        resposta = base_com_instrutor.get(f"/simulacoes/comparar?ids={d1['id']},{d2['id']}")

        assert resposta.status_code == 200
        assert "itens" in resposta.json()


class TestExportacao:
    def test_exporta_simulacao_concluida(
        self, base_com_instrutor: TestClient, cenario_basico: dict
    ) -> None:
        disparo = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario_basico["id"]}
        ).json()

        resposta = base_com_instrutor.get(f"/simulacoes/{disparo['id']}/exportar")

        assert resposta.status_code == 200
        assert resposta.headers["content-type"].startswith("application/vnd.openxmlformats")
        assert f"simulacao_{disparo['id']}_" in resposta.headers["content-disposition"]

    def test_planilha_tem_turmas_e_indicadores(
        self, base_com_instrutor: TestClient, cenario_basico: dict
    ) -> None:
        import io

        from openpyxl import load_workbook

        disparo = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario_basico["id"]}
        ).json()

        conteudo = base_com_instrutor.get(f"/simulacoes/{disparo['id']}/exportar").content
        workbook = load_workbook(io.BytesIO(conteudo))

        assert "Turmas Sugeridas" in workbook.sheetnames
        assert "Indicadores e Cenário" in workbook.sheetnames

        aba_turmas = workbook["Turmas Sugeridas"]
        assert aba_turmas.max_row > 1  # cabeçalho + ao menos uma turma

    def test_recusa_exportar_simulacao_nao_concluida(
        self, base_com_instrutor: TestClient, cenario_basico: dict, monkeypatch
    ) -> None:
        """Simula uma simulação travada em 'executando' interceptando a tarefa
        de background antes que ela rode até a conclusão."""
        import app.api.simulacoes as modulo_simulacoes

        monkeypatch.setattr(modulo_simulacoes, "executar_simulacao", lambda _id: None)

        disparo = base_com_instrutor.post(
            "/simulacoes/executar", json={"cenario_id": cenario_basico["id"]}
        ).json()

        resposta = base_com_instrutor.get(f"/simulacoes/{disparo['id']}/exportar")

        assert resposta.status_code == 422

    def test_exportar_simulacao_inexistente_retorna_404(
        self, base_com_instrutor: TestClient
    ) -> None:
        assert base_com_instrutor.get("/simulacoes/9999/exportar").status_code == 404


class TestConsultas:
    def test_simulacao_inexistente_retorna_404(self, client: TestClient) -> None:
        assert client.get("/simulacoes/9999").status_code == 404
        assert client.get("/simulacoes/9999/turmas-sugeridas").status_code == 404
        assert client.get("/simulacoes/9999/kpis").status_code == 404

    def test_lista_historico(self, base_com_instrutor: TestClient, cenario_basico: dict) -> None:
        base_com_instrutor.post("/simulacoes/executar", json={"cenario_id": cenario_basico["id"]})
        base_com_instrutor.post("/simulacoes/executar", json={"cenario_id": cenario_basico["id"]})

        historico = base_com_instrutor.get("/simulacoes").json()

        assert len(historico) == 2

    def test_filtra_historico_por_cenario(self, base_com_instrutor: TestClient) -> None:
        cenario_a = base_com_instrutor.post(
            "/cenarios",
            json={
                "nome": "A",
                "periodo_de": "2026-08-31",
                "periodo_ate": "2026-11-30",
                "pesos_objetivo": {
                    "maximizar_aproveitamento": 1.0,
                    "antecipar_inicio": 0.0,
                    "balancear_carga_instrutores": 0.0,
                    "balancear_tipologias": 0.0,
                },
            },
        ).json()
        cenario_b = base_com_instrutor.post(
            "/cenarios",
            json={
                "nome": "B",
                "periodo_de": "2026-08-31",
                "periodo_ate": "2026-11-30",
                "pesos_objetivo": {
                    "maximizar_aproveitamento": 1.0,
                    "antecipar_inicio": 0.0,
                    "balancear_carga_instrutores": 0.0,
                    "balancear_tipologias": 0.0,
                },
            },
        ).json()
        base_com_instrutor.post("/simulacoes/executar", json={"cenario_id": cenario_a["id"]})
        base_com_instrutor.post("/simulacoes/executar", json={"cenario_id": cenario_b["id"]})

        historico_a = base_com_instrutor.get(f"/simulacoes?cenario_id={cenario_a['id']}").json()

        assert len(historico_a) == 1
        assert historico_a[0]["cenario_id"] == cenario_a["id"]
