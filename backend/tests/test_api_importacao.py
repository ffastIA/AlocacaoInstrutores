"""Testes dos endpoints de importação e das planilhas-modelo."""

import io

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from tests.fabricas import planilha_instrutores, planilha_tipologias

CSV = "text/csv"


def _enviar(client: TestClient, rota: str, conteudo: bytes, nome: str = "planilha.csv"):
    return client.post(rota, files={"arquivo": (nome, conteudo, CSV)})


class TestUploadInstrutores:
    def test_importacao_bem_sucedida(self, client: TestClient) -> None:
        conteudo = planilha_instrutores(
            [["Maria Silva", "Jovem Digital", "manha", "4", "2;4", "Programação"]]
        )

        resposta = _enviar(client, "/importar/instrutores", conteudo)

        assert resposta.status_code == 200
        corpo = resposta.json()
        assert corpo["importados"] == 1
        assert corpo["erros"] == []

    def test_relatorio_de_erros_por_linha(self, client: TestClient) -> None:
        conteudo = planilha_instrutores(
            [
                ["Maria Silva", "Jovem Digital", "manha", "4", "2;4", "Programação"],
                ["Erro", "Jovem Digital", "manha;tarde", "4", "2;4", "Programação"],
            ]
        )

        corpo = _enviar(client, "/importar/instrutores", conteudo).json()

        assert corpo["importados"] == 1
        assert corpo["rejeitados"] == 1
        assert corpo["erros"][0]["linha"] == 3
        assert "mesma quantidade" in corpo["erros"][0]["motivo"]

    def test_arquivo_sem_coluna_obrigatoria(self, client: TestClient) -> None:
        from tests.fabricas import csv_bytes

        conteudo = csv_bytes(["nome", "projeto"], [["Maria", "Jovem Digital"]])

        corpo = _enviar(client, "/importar/instrutores", conteudo).json()

        assert corpo["sucesso"] is False
        assert corpo["erro_arquivo"] is not None
        assert corpo["importados"] == 0

    def test_arquivo_vazio_e_recusado(self, client: TestClient) -> None:
        resposta = _enviar(client, "/importar/instrutores", b"")

        assert resposta.status_code == 400

    def test_formato_nao_suportado(self, client: TestClient) -> None:
        corpo = _enviar(client, "/importar/instrutores", b"dados", "arquivo.pdf").json()

        assert corpo["erro_arquivo"] is not None
        assert ".xlsx" in corpo["erro_arquivo"]

    def test_alerta_de_tipologias_pendentes(self, client: TestClient) -> None:
        conteudo = planilha_instrutores(
            [["Maria Silva", "Jovem Digital", "manha", "4", "2;4", "Robótica"]]
        )

        corpo = _enviar(client, "/importar/instrutores", conteudo).json()

        assert any("pendente" in a["mensagem"] for a in corpo["alertas"])


class TestUploadTipologias:
    def test_configura_carga_horaria(self, client: TestClient) -> None:
        _enviar(
            client,
            "/importar/instrutores",
            planilha_instrutores(
                [["Maria Silva", "Jovem Digital", "manha", "4", "2;4", "Robótica"]]
            ),
        )

        corpo = _enviar(
            client, "/importar/tipologias", planilha_tipologias([["Robótica", "40", "4"]])
        ).json()

        assert corpo["atualizados"] == 1
        assert client.get("/tipologias/pendentes").json() == []

    def test_rejeita_carga_nao_divisivel(self, client: TestClient) -> None:
        corpo = _enviar(
            client, "/importar/tipologias", planilha_tipologias([["Robótica", "50", "4"]])
        ).json()

        assert "múltiplo exato" in corpo["erros"][0]["motivo"]


class TestModelos:
    def test_lista_tipos_disponiveis(self, client: TestClient) -> None:
        tipos = client.get("/importar/modelos").json()["tipos"]

        assert set(tipos) == {"instrutores", "tipologias", "turmas-em-andamento"}

    def test_baixa_modelo_de_instrutores(self, client: TestClient) -> None:
        resposta = client.get("/importar/modelos/instrutores")

        assert resposta.status_code == 200
        assert "modelo_instrutores.xlsx" in resposta.headers["content-disposition"]

    def test_modelo_tem_cabecalhos_e_exemplo(self, client: TestClient) -> None:
        conteudo = client.get("/importar/modelos/instrutores").content
        aba = load_workbook(io.BytesIO(conteudo))["Dados"]

        cabecalhos = [c.value for c in aba[1]]
        assert "nome" in cabecalhos
        assert "tipologias" in cabecalhos

        exemplo = [c.value for c in aba[2]]
        assert ";" in str(exemplo[cabecalhos.index("turnos")])

    def test_orientacoes_ficam_em_aba_separada(self, client: TestClient) -> None:
        """Se ficassem sob os dados, seriam lidas como registros ao reimportar."""
        conteudo = client.get("/importar/modelos/instrutores").content
        workbook = load_workbook(io.BytesIO(conteudo))

        assert workbook.sheetnames[0] == "Dados"
        assert "Como preencher" in workbook.sheetnames

    def test_modelo_baixado_e_importavel(self, client: TestClient) -> None:
        """O ciclo baixar → importar precisa funcionar sem edição."""
        conteudo = client.get("/importar/modelos/instrutores").content

        resposta = client.post(
            "/importar/instrutores",
            files={
                "arquivo": (
                    "modelo.xlsx",
                    conteudo,
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )

        corpo = resposta.json()
        assert corpo["erro_arquivo"] is None
        assert corpo["erros"] == []
        assert corpo["importados"] == 3

    def test_tipo_desconhecido_retorna_404(self, client: TestClient) -> None:
        resposta = client.get("/importar/modelos/inexistente")

        assert resposta.status_code == 404
        assert "instrutores" in resposta.json()["detail"]
