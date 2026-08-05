"""Leitura de planilhas `.xlsx` e `.csv`.

Os cabeçalhos são normalizados na leitura, para que a equipe não precise
padronizar a planilha antes do primeiro uso: `Dias Semana`, `dias_semana` e
`DIAS SEMANA` são equivalentes.
"""

import csv
import io
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook

from app.services.importacao.resultado import ArquivoInvalidoError

# Primeira linha de dados numa planilha com cabeçalho.
PRIMEIRA_LINHA_DADOS = 2


def normalizar_cabecalho(texto: str) -> str:
    """Reduz um cabeçalho à sua forma canônica.

    Remove acentos, converte para minúsculo e troca espaços e hifens por
    underscore: `"Carga Horária Turno"` vira `"carga_horaria_turno"`.
    """
    sem_acento = "".join(
        c for c in unicodedata.normalize("NFD", texto) if unicodedata.category(c) != "Mn"
    )
    canonico = sem_acento.strip().lower()
    for separador in (" ", "-", "."):
        canonico = canonico.replace(separador, "_")
    while "__" in canonico:
        canonico = canonico.replace("__", "_")
    return canonico.strip("_")


@dataclass
class Linha:
    """Uma linha de dados, indexada por cabeçalho normalizado.

    `numero` é a posição na planilha (1-based, contando o cabeçalho), para que
    a mensagem de erro aponte a linha que o usuário vê no Excel.
    """

    numero: int
    valores: dict[str, Any]

    def texto(self, coluna: str) -> str:
        """Valor como texto, já sem espaços nas extremidades. Vazio se ausente."""
        valor = self.valores.get(coluna)
        if valor is None:
            return ""
        if isinstance(valor, datetime):
            return valor.date().isoformat()
        if isinstance(valor, date):
            return valor.isoformat()
        if isinstance(valor, float) and valor.is_integer():
            return str(int(valor))
        return str(valor).strip()

    def bruto(self, coluna: str) -> Any:
        return self.valores.get(coluna)

    def vazia(self) -> bool:
        return all(not str(v).strip() for v in self.valores.values() if v is not None)


@dataclass
class Planilha:
    cabecalhos: list[str]
    linhas: list[Linha]

    def exigir_colunas(self, obrigatorias: list[str]) -> None:
        """Recusa o arquivo inteiro se faltar coluna obrigatória.

        É falha estrutural: sem a coluna, nenhuma linha pode ser validada.
        """
        faltando = [c for c in obrigatorias if c not in self.cabecalhos]
        if faltando:
            raise ArquivoInvalidoError(
                "Coluna obrigatória ausente: "
                + ", ".join(f"'{c}'" for c in faltando)
                + f". Colunas encontradas: {', '.join(self.cabecalhos) or '(nenhuma)'}."
            )

    def tem_coluna(self, coluna: str) -> bool:
        return coluna in self.cabecalhos


def ler_planilha(conteudo: bytes, nome_arquivo: str) -> Planilha:
    """Lê o conteúdo de uma planilha, escolhendo o parser pela extensão."""
    nome = nome_arquivo.lower()
    if nome.endswith(".xlsx") or nome.endswith(".xlsm"):
        return _ler_xlsx(conteudo)
    if nome.endswith(".csv"):
        return _ler_csv(conteudo)
    raise ArquivoInvalidoError(
        f"Formato não suportado: '{nome_arquivo}'. Envie um arquivo .xlsx ou .csv."
    )


def _ler_xlsx(conteudo: bytes) -> Planilha:
    try:
        # data_only lê o resultado das fórmulas, não a fórmula em si.
        workbook = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - qualquer falha aqui é arquivo ilegível
        raise ArquivoInvalidoError(f"Não foi possível ler a planilha: {exc}") from exc

    try:
        aba = workbook.active
        if aba is None:
            raise ArquivoInvalidoError("A planilha não contém nenhuma aba.")

        iterador = aba.iter_rows(values_only=True)
        try:
            cabecalho_bruto = next(iterador)
        except StopIteration:
            raise ArquivoInvalidoError("A planilha está vazia.") from None

        cabecalhos = [normalizar_cabecalho(str(c)) for c in cabecalho_bruto if c is not None]
        linhas = [
            linha
            for indice, valores in enumerate(iterador, start=PRIMEIRA_LINHA_DADOS)
            if not (linha := Linha(numero=indice, valores=_zipar(cabecalhos, valores))).vazia()
        ]
        return Planilha(cabecalhos=cabecalhos, linhas=linhas)
    finally:
        workbook.close()


def _ler_csv(conteudo: bytes) -> Planilha:
    texto = _decodificar(conteudo)
    # A planilha usa ';' dentro das células (listas de turnos, tipologias), então
    # o separador de colunas do CSV precisa ser ',' — detectar automaticamente
    # confundiria os dois papéis do ponto e vírgula.
    leitor = csv.DictReader(io.StringIO(texto), delimiter=",")

    if leitor.fieldnames is None:
        raise ArquivoInvalidoError("O arquivo CSV está vazio.")

    cabecalhos = [normalizar_cabecalho(c) for c in leitor.fieldnames if c]
    linhas: list[Linha] = []
    for indice, registro in enumerate(leitor, start=PRIMEIRA_LINHA_DADOS):
        valores = {
            normalizar_cabecalho(chave): valor
            for chave, valor in registro.items()
            if chave is not None
        }
        linha = Linha(numero=indice, valores=valores)
        if not linha.vazia():
            linhas.append(linha)

    return Planilha(cabecalhos=cabecalhos, linhas=linhas)


def _decodificar(conteudo: bytes) -> str:
    """Decodifica tentando UTF-8 e caindo para Latin-1.

    Planilhas exportadas do Excel em português costumam vir em Latin-1, e um
    acento mal decodificado corromperia nomes de instrutor e de tipologia.
    """
    for codificacao in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return conteudo.decode(codificacao)
        except UnicodeDecodeError:
            continue
    raise ArquivoInvalidoError("Não foi possível decodificar o arquivo.")


def _zipar(cabecalhos: list[str], valores: tuple) -> dict[str, Any]:
    return {cabecalho: valores[i] if i < len(valores) else None
            for i, cabecalho in enumerate(cabecalhos)}
