"""Geração das planilhas-modelo para download.

Cada modelo traz uma linha de exemplo preenchida, demonstrando o formato dos
campos multivalorados — o usuário não precisa consultar documentação externa
para saber que o separador é ponto e vírgula.
"""

import io
from dataclasses import dataclass

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass
class ModeloPlanilha:
    nome_arquivo: str
    cabecalhos: list[str]
    exemplos: list[list[str]]


MODELOS: dict[str, ModeloPlanilha] = {
    "instrutores": ModeloPlanilha(
        nome_arquivo="modelo_instrutores.xlsx",
        cabecalhos=[
            "nome",
            "projeto",
            "turnos",
            "dias_semana",
            "tipologias",
            "observacao",
        ],
        exemplos=[
            [
                "Maria Silva",
                "Jovem Digital",
                "manha_1;manha_2;tarde_1",
                "2;3;4;5",
                "Programação;Pixel Art",
                "Prefere turno da manhã",
            ],
            ["João Souza", "Jovem Digital", "noite", "2;4", "Robótica", ""],
            [
                "Ana Costa",
                "Inclusão Tech",
                "manha_1;noite",
                "2;3;4;5;6",
                "Robótica;Programação;Google Workspace",
                "",
            ],
        ],
    ),
    "tipologias": ModeloPlanilha(
        nome_arquivo="modelo_tipologias.xlsx",
        cabecalhos=["tipologia", "carga_horaria_total", "horas_por_encontro", "descricao"],
        exemplos=[
            ["Programação", "60", "3", ""],
            ["Pixel Art", "24", "2", ""],
            ["Robótica", "40", "4", "Montagem e programação de kits"],
        ],
    ),
    "turmas-em-andamento": ModeloPlanilha(
        nome_arquivo="modelo_turmas_em_andamento.xlsx",
        cabecalhos=[
            "instrutor",
            "tipologia",
            "modalidade",
            "turno",
            "data_inicio",
            "data_fim_prevista",
            "codigo_turma",
        ],
        exemplos=[
            [
                "Maria Silva",
                "Programação",
                "regular_seg_qua",
                "manha_1",
                "01/06/2026",
                "30/08/2026",
                "PROG-2026-014",
            ],
            [
                "João Souza",
                "Robótica",
                "intensiva_seg_qui",
                "noite",
                "15/07/2026",
                "20/09/2026",
                "",
            ],
        ],
    ),
    "datas-nao-letivas": ModeloPlanilha(
        nome_arquivo="modelo_datas_nao_letivas.xlsx",
        cabecalhos=["data_inicio", "data_fim", "descricao", "tipo", "projeto"],
        exemplos=[
            ["07/09/2026", "", "Independência", "feriado", ""],
            ["24/12/2026", "06/01/2027", "Recesso de fim de ano", "recesso", ""],
            ["15/03/2027", "19/03/2027", "Férias da equipe", "ferias", "Inclusão Tech"],
        ],
    ),
}

# Orientações de preenchimento. Ficam numa **aba separada**: se estivessem
# abaixo dos dados, seriam lidas como registros ao reimportar o arquivo.
ORIENTACOES: dict[str, list[str]] = {
    "instrutores": [
        "Campos com várias opções usam ponto e vírgula como separador.",
        "turnos: manha_1, manha_2, tarde_1, tarde_2 ou noite. Ex.: manha_1;tarde_1",
        "  Manhã e tarde têm dois horários fixos cada (slots 1 e 2); a noite tem só um.",
        "  Cada slot comporta no máximo uma turma por vez.",
        "dias_semana: 2=segunda, 3=terça, 4=quarta, 5=quinta, 6=sexta. Ex.: 2;4",
        "  Sexta (6) conta apenas como capacidade de reposição, nunca recebe turma regular.",
        "tipologias: nomes dos cursos que o instrutor domina. Ex.: Programação;Pixel Art",
    ],
    "tipologias": [
        "carga_horaria_total: entre 24 e 60 horas.",
        "horas_por_encontro: precisa dividir a carga total em valor inteiro.",
        "  Ex.: 40h com 4h por encontro = 10 encontros (válido).",
        "  Ex.: 50h com 4h por encontro = 12,5 encontros (inválido).",
    ],
    "turmas-em-andamento": [
        "modalidade: regular_seg_qua, regular_ter_qui ou intensiva_seg_qui.",
        "turno: manha_1, manha_2, tarde_1, tarde_2 ou noite. "
        "Precisa constar na disponibilidade do instrutor.",
        "Datas no formato DD/MM/AAAA.",
        "Deixe a planilha sem linhas de dados se nenhuma turma estiver em curso.",
    ],
    "datas-nao-letivas": [
        "Datas no formato DD/MM/AAAA.",
        "data_fim: deixe vazio para um intervalo de um único dia.",
        "tipo: feriado, recesso ou ferias. Vazio equivale a 'feriado'.",
        "projeto: nome do projeto ao qual o intervalo se aplica. Vazio aplica a todos.",
        "Estes dados ainda não afetam o cálculo das simulações nesta versão.",
    ],
}


class ModeloDesconhecidoError(ValueError):
    """Tipo de modelo solicitado não existe."""


def tipos_disponiveis() -> list[str]:
    return sorted(MODELOS)


def gerar_modelo(tipo: str) -> tuple[bytes, str]:
    """Gera a planilha-modelo do tipo pedido, retornando conteúdo e nome."""
    modelo = MODELOS.get(tipo)
    if modelo is None:
        raise ModeloDesconhecidoError(
            f"Modelo desconhecido: '{tipo}'. Tipos disponíveis: {', '.join(tipos_disponiveis())}"
        )

    workbook = Workbook()
    aba = workbook.active
    aba.title = "Dados"

    aba.append(modelo.cabecalhos)
    for celula in aba[1]:
        celula.font = Font(bold=True)
        celula.fill = PatternFill("solid", start_color="DDEBF7")
        celula.alignment = Alignment(horizontal="center")

    for exemplo in modelo.exemplos:
        aba.append(exemplo)

    _ajustar_larguras(aba, modelo)
    _criar_aba_orientacoes(workbook, tipo)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue(), modelo.nome_arquivo


def _ajustar_larguras(aba, modelo: ModeloPlanilha) -> None:
    for indice, cabecalho in enumerate(modelo.cabecalhos, start=1):
        conteudos = [cabecalho] + [
            str(linha[indice - 1]) for linha in modelo.exemplos if indice <= len(linha)
        ]
        largura = min(max(len(c) for c in conteudos) + 4, 45)
        aba.column_dimensions[get_column_letter(indice)].width = largura


def _criar_aba_orientacoes(workbook: Workbook, tipo: str) -> None:
    """Cria a aba de instruções, separada da aba de dados.

    A separação importa: o leitor de planilha processa apenas a primeira aba,
    então as orientações nunca são confundidas com registros.
    """
    orientacoes = ORIENTACOES.get(tipo, [])
    if not orientacoes:
        return

    aba = workbook.create_sheet("Como preencher")
    aba.column_dimensions["A"].width = 90

    titulo = aba.cell(row=1, column=1, value="COMO PREENCHER")
    titulo.font = Font(bold=True, size=12)

    for deslocamento, texto in enumerate(orientacoes, start=3):
        aba.cell(row=deslocamento, column=1, value=texto)

    nota = aba.cell(
        row=len(orientacoes) + 5,
        column=1,
        value=(
            "Substitua as linhas de exemplo da aba 'Dados' pelos seus registros antes de importar."
        ),
    )
    nota.font = Font(bold=True, italic=True)
