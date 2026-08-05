"""Exportação do resultado de uma simulação em planilha."""

import io
import re
import unicodedata

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.models import ResultadoKpis, Simulacao, TurmaSugerida
from app.services.cenarios.parametros import carregar_parametros

CABECALHOS_TURMAS = [
    "tipologia",
    "instrutor",
    "projeto",
    "modalidade",
    "turno",
    "data_inicio",
    "data_fim",
    "num_encontros",
    "carga_horaria_total",
]


def gerar_planilha_resultado(db: Session, simulacao: Simulacao) -> tuple[bytes, str]:
    """Gera a planilha com turmas sugeridas, KPIs e parâmetros do cenário."""
    workbook = Workbook()

    _aba_turmas(workbook, db, simulacao)
    _aba_indicadores(workbook, simulacao)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue(), _nome_arquivo(simulacao)


def _aba_turmas(workbook: Workbook, db: Session, simulacao: Simulacao) -> None:
    aba = workbook.active
    aba.title = "Turmas Sugeridas"

    aba.append(CABECALHOS_TURMAS)
    for celula in aba[1]:
        celula.font = Font(bold=True)
        celula.fill = PatternFill("solid", start_color="DDEBF7")

    turmas = db.scalars(
        select(TurmaSugerida)
        .where(TurmaSugerida.simulacao_id == simulacao.id)
        .options(
            selectinload(TurmaSugerida.tipologia),
            selectinload(TurmaSugerida.instrutor),
            selectinload(TurmaSugerida.projeto),
        )
        .order_by(TurmaSugerida.data_inicio)
    ).all()

    for t in turmas:
        aba.append(
            [
                t.tipologia.nome,
                t.instrutor.nome,
                t.projeto.nome,
                t.modalidade.value,
                t.turno.value,
                t.data_inicio.isoformat(),
                t.data_fim.isoformat(),
                t.num_encontros,
                t.carga_horaria_total,
            ]
        )

    for indice in range(1, len(CABECALHOS_TURMAS) + 1):
        aba.column_dimensions[get_column_letter(indice)].width = 20


def _aba_indicadores(workbook: Workbook, simulacao: Simulacao) -> None:
    aba = workbook.create_sheet("Indicadores e Cenário")
    aba.column_dimensions["A"].width = 32
    aba.column_dimensions["B"].width = 30

    def _linha(rotulo: str, valor: object) -> None:
        aba.append([rotulo, valor])

    titulo = aba.cell(row=1, column=1, value="INDICADORES")
    titulo.font = Font(bold=True, size=12)
    aba.append([])

    kpis: ResultadoKpis | None = simulacao.kpis
    if kpis is not None:
        _linha("Total de turmas sugeridas", kpis.total_turmas_sugeridas)
        _linha("Horas de formação total", kpis.horas_formacao_total)
        _linha("Ociosidade (%)", round(kpis.pct_ociosidade, 1))
        _linha("Índice de balanceamento de carga", round(kpis.indice_balanceamento_carga, 1))
        _linha("Índice de balanceamento de tipologias", kpis.indice_balanceamento_tipologia)
        _linha("Horas de reposição (sextas)", kpis.horas_reposicao_sexta)

    aba.append([])
    subtitulo = aba.cell(row=aba.max_row + 1, column=1, value="PARÂMETROS DO CENÁRIO")
    subtitulo.font = Font(bold=True, size=12)
    aba.append([])

    _linha("Cenário", simulacao.cenario.nome)
    _linha("Período — de", simulacao.cenario.periodo_de.isoformat())
    _linha("Período — até", simulacao.cenario.periodo_ate.isoformat())
    _linha(
        "Compartilhamento entre projetos",
        "Sim" if simulacao.cenario.permitir_compartilhamento else "Não",
    )

    parametros = carregar_parametros(simulacao.parametros_json_path)
    aba.append([])
    _linha("Peso — maximizar aproveitamento", parametros.pesos_objetivo.maximizar_aproveitamento)
    _linha("Peso — antecipar início", parametros.pesos_objetivo.antecipar_inicio)
    _linha("Peso — balancear carga", parametros.pesos_objetivo.balancear_carga_instrutores)
    _linha("Peso — balancear tipologias", parametros.pesos_objetivo.balancear_tipologias)


def _nome_arquivo(simulacao: Simulacao) -> str:
    nome_cenario = _slug(simulacao.cenario.nome)
    data_execucao = (simulacao.concluido_em or simulacao.iniciado_em).strftime("%Y%m%d_%H%M")
    return f"simulacao_{simulacao.id}_{nome_cenario}_{data_execucao}.xlsx"


def _slug(texto: str) -> str:
    sem_acento = "".join(
        c for c in unicodedata.normalize("NFD", texto) if unicodedata.category(c) != "Mn"
    )
    return re.sub(r"[^a-zA-Z0-9]+", "_", sem_acento).strip("_").lower()
